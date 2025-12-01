import calendar
from datetime import datetime, date
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from typing import Dict, List, Any
import logging

# --- STILLAR VA FORMATLAR ---
class ExcelStyles:
    # Ranglar
    HEADER_FILL = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    BLOCK_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    ABSENT_FILL = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
    POSITIVE_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    WARNING_FILL = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
    
    # Shriftlar
    TITLE_FONT = Font(bold=True, name='Arial', size=14)
    HEADER_FONT = Font(bold=True, name='Arial', size=11)
    DATA_FONT = Font(name='Arial', size=10)
    BOLD_FONT = Font(bold=True, name='Arial', size=10)
    
    # Joylashuv
    CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT_ALIGN = Alignment(horizontal="left", vertical="center")
    RIGHT_ALIGN = Alignment(horizontal="right", vertical="center")
    
    # Chegaralar
    THIN_BORDER = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

# O'zbekcha oy nomlari
MONTHS_UZ = {
    1: "YANVAR", 2: "FEVRAL", 3: "MART", 4: "APREL", 5: "MAY", 6: "IYUN",
    7: "IYUL", 8: "AVGUST", 9: "SENTABR", 10: "OKTABR", 11: "NOYABR", 12: "DEKABR"
}

def generate_report(year: int, month: int, workers_data: List[Dict], 
                   attendance_data: Dict, advances_data: Dict) -> str:
    """
    Excel hisobot yaratish
    """
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = f"{month:02d}-{year}"
        
        # Ma'lumotlarni saralash (faqat ism bo'yicha)
        workers_data.sort(key=lambda x: x['name'])
        
        num_days = calendar.monthrange(year, month)[1]
        styles = ExcelStyles()
        
        # SARLAVHA YARATISH
        _create_header(ws, year, month, num_days, styles)
        
        # ISHCHI MA'LUMOTLARINI QO'SHISH
        _add_worker_data(ws, workers_data, attendance_data, advances_data, 
                        year, month, num_days, styles)
        
        # AUTOSIZE USTUNLAR
        _auto_adjust_columns(ws)
        
        # FAYLNI SAQLASH
        filename = f"hisobot_{MONTHS_UZ.get(month, month)}_{year}.xlsx"
        wb.save(filename)
        logging.info(f"✅ Excel hisobot yaratildi: {filename}")
        return filename
        
    except Exception as e:
        logging.error(f"❌ Excel hisobot yaratishda xato: {e}")
        raise

def _create_header(ws, year, month, num_days, styles):
    """Sarlavha va ustunlarni yaratish"""
    # Asosiy sarlavha
    month_name = MONTHS_UZ.get(month, f"OY-{month}")
    total_cols = 2 + num_days + 5  # № + Ism + Kunlar + Hisob ustunlari
    
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    title_cell = ws.cell(1, 1, f"{month_name} {year} - DAVOMAT VA HISOBOT")
    title_cell.font = styles.TITLE_FONT
    title_cell.alignment = styles.CENTER_ALIGN
    title_cell.fill = styles.HEADER_FILL
    
    # Ustun sarlavhalari
    headers = ["№", "F.I.O"] + [str(day) for day in range(1, num_days + 1)] + [
        "Soatlik narx", "Jami soat", "Avans", "Hisoblangan", "Qo'lga tegadi"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(2, col, header)
        cell.font = styles.HEADER_FONT
        cell.fill = styles.HEADER_FILL
        cell.alignment = styles.CENTER_ALIGN
        cell.border = styles.THIN_BORDER
        
        # Ustun kengliklari
        if col == 1:  # №
            ws.column_dimensions[get_column_letter(col)].width = 6
        elif col == 2:  # F.I.O
            ws.column_dimensions[get_column_letter(col)].width = 35
        elif 3 <= col <= num_days + 2:  # Kunlar
            ws.column_dimensions[get_column_letter(col)].width = 5
        else:  # Hisob ustunlari
            ws.column_dimensions[get_column_letter(col)].width = 15

def _add_worker_data(ws, workers_data, attendance_data, advances_data, 
                    year, month, num_days, styles):
    """Ishchi ma'lumotlarini qo'shish"""
    current_row = 3
    
    for idx, worker in enumerate(workers_data, 1):
        _add_worker_row(ws, current_row, idx, worker, attendance_data, 
                       advances_data, year, month, num_days, styles)
        current_row += 1

def _add_worker_row(ws, row, counter, worker, attendance_data, advances_data,
                   year, month, num_days, styles):
    """Ishchi qatorini qo'shish"""
    # № va Ism
    cell_num = ws.cell(row, 1, counter)
    cell_num.border = styles.THIN_BORDER
    cell_num.alignment = styles.CENTER_ALIGN
    
    cell_name = ws.cell(row, 2, worker['name'])
    cell_name.border = styles.THIN_BORDER
    cell_name.alignment = styles.LEFT_ALIGN
    cell_name.font = styles.DATA_FONT
    
    # DAVOMAT KUNLARI
    total_hours = 0
    
    # created_at va archived_at ni xavfsiz olish
    created_at = worker.get('created_at')
    archived_at = worker.get('archived_at')
    
    # Agar created_at bo'lmasa, oyning birinchi kuni deb olamiz
    if created_at is None:
        created_at = date(year, month, 1)
    elif isinstance(created_at, datetime):
        created_at = created_at.date()
    
    # Agar archived_at bo'lmasa, None qoldiramiz
    if archived_at and isinstance(archived_at, datetime):
        archived_at = archived_at.date()
    
    for day in range(1, num_days + 1):
        col = day + 2
        current_date = date(year, month, day)
        date_key = f"{year}-{month:02d}-{day:02d}"
        
        cell = ws.cell(row, col)
        cell.border = styles.THIN_BORDER
        cell.alignment = styles.CENTER_ALIGN
        
        # Ishlamagan kunlarni belgilash (ishga kirmasdan oldin yoki ketgandan keyin)
        if (current_date < created_at) or (archived_at and current_date > archived_at):
            cell.fill = styles.ABSENT_FILL
        else:
            # attendance_data kaliti (worker_id, 'YYYY-MM-DD') formatida
            hours = attendance_data.get((worker['id'], date_key), 0)
            if hours > 0:
                cell.value = hours
                cell.font = styles.DATA_FONT
                total_hours += hours
            else:
                # Agar ish kuni bo'lsa-yu soat bo'lmasa, bo'sh qoladi yoki 0
                pass
    
    # HISOB-KITOBlAR
    start_calc = num_days + 3
    rate = float(worker['rate'])
    advance = advances_data.get(worker['id'], 0)
    calculated = total_hours * rate
    net_amount = calculated - advance
    
    # Soatlik narx
    c_rate = ws.cell(row, start_calc, rate)
    c_rate.border = styles.THIN_BORDER
    c_rate.number_format = '#,##0'
    
    # Jami soat
    c_total = ws.cell(row, start_calc + 1, total_hours)
    c_total.border = styles.THIN_BORDER
    c_total.font = styles.BOLD_FONT
    c_total.alignment = styles.CENTER_ALIGN
    
    # Avans
    c_adv = ws.cell(row, start_calc + 2, advance)
    c_adv.border = styles.THIN_BORDER
    c_adv.number_format = '#,##0'
    
    # Hisoblangan
    c_calc = ws.cell(row, start_calc + 3, calculated)
    c_calc.border = styles.THIN_BORDER
    c_calc.number_format = '#,##0'
    
    # Qo'lga tegadi
    c_net = ws.cell(row, start_calc + 4, net_amount)
    c_net.border = styles.THIN_BORDER
    c_net.font = styles.BOLD_FONT
    c_net.number_format = '#,##0'
    c_net.fill = styles.POSITIVE_FILL if net_amount >= 0 else styles.WARNING_FILL

def _auto_adjust_columns(ws):
    """Ustunlarni avtomatik sozlash"""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
